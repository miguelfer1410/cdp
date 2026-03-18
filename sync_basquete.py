"""
sync_basquete.py
────────────────────────────────────────────────────────────────────
Sincroniza os atletas de Basquetebol (MINIS + BSQ) do Excel com a BD.

Lógica por atleta (linha do Excel):
  ┌─ Procura na BD pelo MembershipNumber (Nº Sócio do Excel → "CDP-XXXXX")
  │
  ├─ ENCONTRADO
  │     ├─ Verifica se tem AthleteProfile  → cria se não tiver
  │     └─ Verifica se está na equipa certa → adiciona (AthleteTeam) se não estiver
  │
  └─ NÃO ENCONTRADO pelo MembershipNumber
        ├─ Tenta encontrar por nome normalizado
        │     ├─ ENCONTRADO por nome
        │     │     ├─ Atualiza MembershipNumber para o valor do Excel
        │     │     ├─ Verifica/cria AthleteProfile
        │     │     └─ Verifica/cria atribuição de equipa
        │     └─ NÃO ENCONTRADO → regista como "não encontrado" (sem tocar na BD)
        └─ (sem match)

Ligação: SQL Server · DESKTOP-KU8TIMC\MSSQLSERVERS · base de dados: cdp
         Autenticação Windows (Trusted_Connection=yes)
"""

import sys, re, unicodedata
from datetime import datetime

# ── dependências ──────────────────────────────────────────────────
def ensure(pkg):
    try:
        __import__(pkg)
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])

ensure("pyodbc")
ensure("openpyxl")

import pyodbc
import openpyxl

# ─────────────────────────────────────────────────────────────────
# CONFIGURAÇÃO  ← ajusta aqui se necessário
# ─────────────────────────────────────────────────────────────────
EXCEL_PATH = r"ANÁLISE BASQUETE 25-26.xlsx"

CONNECTION_STRING = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    r"SERVER=DESKTOP-KU8TIMC\MSSQLSERVERS;"
    "DATABASE=cdp;"
    "Trusted_Connection=yes;"
)

SHEETS = ["MINIS", "BSQ"]

# Candidatos de nome (minúsculas / sem acentos) a pesquisar na BD
# para cada nome de equipa do Excel.
TEAM_NAME_CANDIDATES = {
    "SUB 8":    ["sub 8",  "sub8",  "mini sub 8",  "minis sub 8"],
    "SUB 10":   ["sub 10", "sub10", "mini sub 10", "minis sub 10"],
    "SUB 12":   ["sub 12", "sub12", "mini sub 12", "minis sub 12"],
    "SUB 13 F": ["sub 13 f", "sub13f", "sub 13 feminino", "sub 13f"],
    "SUB 13 M": ["sub 13 m", "sub13m", "sub 13 masculino", "sub 13m", "sub 13"],
    "SUB 14 F": ["sub 14 f", "sub14f", "sub 14 feminino", "sub 14f"],
    "SUB 14 M": ["sub 14 m", "sub14m", "sub 14 masculino", "sub 14m", "sub 14"],
    "SUB 15 M": ["sub 15 m", "sub15m", "sub 15 masculino", "sub 15m", "sub 15"],
    "SUB 16 F": ["sub 16 f", "sub16f", "sub 16 feminino", "sub 16f"],
    "SUB 16 M": ["sub 16 m", "sub16m", "sub 16 masculino", "sub 16m", "sub 16"],
    "SUB 18 F": ["sub 18 f", "sub18f", "sub 18 feminino", "sub 18f"],
    "SUB 18 M": ["sub 18 m", "sub18m", "sub 18 masculino", "sub 18m", "sub 18"],
}

# ─────────────────────────────────────────────────────────────────
# UTILITÁRIOS
# ─────────────────────────────────────────────────────────────────
def normalize(text: str) -> str:
    """Remove acentos, minúsculas, colapsa espaços."""
    nfkd = unicodedata.normalize("NFKD", str(text))
    return re.sub(r"\s+", " ", nfkd.encode("ascii", "ignore").decode().lower().strip())


def socio_to_membership(n: int) -> str:
    """Converte número inteiro do Excel para formato da BD: 7949 → '07949'."""
    return f"{n:05d}"


# ─────────────────────────────────────────────────────────────────
# LER EXCEL
# ─────────────────────────────────────────────────────────────────
def read_excel(path: str) -> list[dict]:
    """
    Devolve lista de dicts:
      { name, socio (int), membership_number (str), team_excel (str), sheet (str) }
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    athletes = []

    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"[AVISO] Folha '{sheet_name}' não encontrada no Excel.")
            continue

        ws = wb[sheet_name]
        current_team = None

        for row in ws.iter_rows(values_only=True):
            col_a = row[0] if len(row) > 0 else None
            col_b = row[1] if len(row) > 1 else None
            col_c = row[2] if len(row) > 2 else None

            # Cabeçalho de equipa: col A = nome da equipa, col B = "Nome"
            if col_a and isinstance(col_a, str) and col_b == "Nome":
                current_team = col_a.strip()
                continue

            # Linha de atleta: col B = nome, col C = nº sócio numérico
            if (current_team
                    and col_b and isinstance(col_b, str) and col_b.strip()
                    and col_c and isinstance(col_c, (int, float))):
                socio = int(col_c)
                athletes.append({
                    "name":              col_b.strip(),
                    "socio":             socio,
                    "membership_number": socio_to_membership(socio),
                    "team_excel":        current_team,
                    "sheet":             sheet_name,
                })

    print(f"\n✔ Excel lido: {len(athletes)} atletas em {len(SHEETS)} folhas.\n")
    return athletes


# ─────────────────────────────────────────────────────────────────
# BASE DE DADOS — funções auxiliares
# ─────────────────────────────────────────────────────────────────
def get_all_teams(cursor) -> dict[str, int]:
    """Devolve { normalize(team_name): team_id } para todas as equipas."""
    cursor.execute("SELECT Id, Name FROM Teams")
    return {normalize(row.Name): row.Id for row in cursor.fetchall()}


def resolve_team_id(team_excel: str, all_teams: dict[str, int]) -> int | None:
    """
    Tenta fazer match do nome de equipa Excel com um Id na BD.
    Usa TEAM_NAME_CANDIDATES para variantes e depois substring match.
    """
    candidates = TEAM_NAME_CANDIDATES.get(team_excel, [normalize(team_excel)])

    # 1. Match exato por candidatos
    for c in candidates:
        if c in all_teams:
            return all_teams[c]

    # 2. Substring: algum candidato contido no nome da BD
    for c in candidates:
        for db_name, tid in all_teams.items():
            if c in db_name or db_name in c:
                return tid

    return None


def find_user_by_membership(cursor, membership_number: str) -> dict | None:
    """Procura utilizador + perfis pelo MembershipNumber."""
    cursor.execute("""
        SELECT
            u.Id         AS UserId,
            u.FirstName,
            u.LastName,
            mp.Id        AS MemberProfileId,
            mp.MembershipNumber,
            ap.Id        AS AthleteProfileId
        FROM Users u
        JOIN MemberProfiles mp ON mp.UserId = u.Id
        LEFT JOIN AthleteProfiles ap ON ap.UserId = u.Id
        WHERE mp.MembershipNumber = ?
    """, membership_number)
    row = cursor.fetchone()
    if not row:
        return None
    return {
        "user_id":           row.UserId,
        "first_name":        row.FirstName,
        "last_name":         row.LastName,
        "member_profile_id": row.MemberProfileId,
        "membership_number": row.MembershipNumber,
        "athlete_profile_id": row.AthleteProfileId,
    }


def find_user_by_name(cursor, name: str) -> dict | None:
    """
    Procura utilizador cujo FirstName + LastName normalizado coincide
    com o nome normalizado do Excel.
    """
    norm_excel = normalize(name)

    cursor.execute("""
        SELECT
            u.Id         AS UserId,
            u.FirstName,
            u.LastName,
            mp.Id        AS MemberProfileId,
            mp.MembershipNumber,
            ap.Id        AS AthleteProfileId
        FROM Users u
        LEFT JOIN MemberProfiles mp ON mp.UserId = u.Id
        LEFT JOIN AthleteProfiles ap ON ap.UserId = u.Id
    """)
    rows = cursor.fetchall()

    for row in rows:
        full_db = normalize(f"{row.FirstName} {row.LastName}")
        if full_db == norm_excel:
            return {
                "user_id":            row.UserId,
                "first_name":         row.FirstName,
                "last_name":          row.LastName,
                "member_profile_id":  row.MemberProfileId,
                "membership_number":  row.MembershipNumber,
                "athlete_profile_id": row.AthleteProfileId,
            }
    return None


def is_athlete_in_team(cursor, athlete_profile_id: int, team_id: int) -> bool:
    cursor.execute("""
        SELECT 1 FROM AthleteTeams
        WHERE AthleteProfileId = ? AND TeamId = ? AND LeftAt IS NULL
    """, athlete_profile_id, team_id)
    return cursor.fetchone() is not None


def create_athlete_profile(cursor, conn, user_id: int) -> int:
    """Cria AthleteProfile e devolve o novo Id."""
    cursor.execute("""
        INSERT INTO AthleteProfiles (UserId, CreatedAt)
        OUTPUT INSERTED.Id
        VALUES (?, GETUTCDATE())
    """, user_id)
    new_id = cursor.fetchone()[0]
    conn.commit()
    return new_id


def assign_athlete_to_team(cursor, conn, athlete_profile_id: int, team_id: int):
    cursor.execute("""
        INSERT INTO AthleteTeams (AthleteProfileId, TeamId, JoinedAt, IsCaptain)
        VALUES (?, ?, GETUTCDATE(), 0)
    """, athlete_profile_id, team_id)
    conn.commit()


def update_membership_number(cursor, conn, member_profile_id: int, new_number: str):
    cursor.execute("""
        UPDATE MemberProfiles SET MembershipNumber = ? WHERE Id = ?
    """, new_number, member_profile_id)
    conn.commit()


# ─────────────────────────────────────────────────────────────────
# LÓGICA PRINCIPAL
# ─────────────────────────────────────────────────────────────────
def process(athletes: list[dict], conn):
    cursor = conn.cursor()
    all_teams = get_all_teams(cursor)

    print(f"Equipas encontradas na BD ({len(all_teams)}):")
    for name in sorted(all_teams):
        print(f"  [{all_teams[name]}] {name}")
    print()

    # Contadores
    stats = {
        "membership_ok":         0,  # encontrado por membership, equipa já ok
        "team_added":            0,  # adicionado à equipa
        "athlete_profile_created": 0,
        "membership_corrected":  0,  # membership corrigido (encontrado por nome)
        "not_found":             0,  # não encontrado de nenhuma forma
        "team_not_found":        0,  # equipa não resolvida na BD
    }
    not_found_list   = []
    team_unresolved  = []
    actions_log      = []

    for ath in athletes:
        name         = ath["name"]
        membership   = ath["membership_number"]
        team_excel   = ath["team_excel"]

        # ── 1. Resolver equipa na BD ──────────────────────────────
        team_id = resolve_team_id(team_excel, all_teams)
        if team_id is None:
            msg = f"[EQUIPA NÃO RESOLVIDA] '{name}' (sócio {ath['socio']}) → equipa Excel '{team_excel}'"
            print(f"  ⚠  {msg}")
            team_unresolved.append({"name": name, "socio": ath["socio"], "team_excel": team_excel})
            stats["team_not_found"] += 1
            continue

        # ── 2. Procurar utilizador pelo MembershipNumber ──────────
        user = find_user_by_membership(cursor, membership)

        if user is None:
            # ── 2b. Tentar por nome ───────────────────────────────
            user = find_user_by_name(cursor, name)

            if user is None:
                msg = f"[NÃO ENCONTRADO] '{name}' (sócio {ath['socio']}, equipa '{team_excel}')"
                print(f"  ✘  {msg}")
                not_found_list.append({"name": name, "socio": ath["socio"], "team_excel": team_excel})
                stats["not_found"] += 1
                continue

            # Encontrado por nome — verificar/corrigir MembershipNumber
            old_mn = user["membership_number"] or "(sem membership)"
            if user["member_profile_id"] and old_mn != membership:
                update_membership_number(cursor, conn, user["member_profile_id"], membership)
                msg = (f"[MEMBERSHIP CORRIGIDO] '{name}' (UserId {user['user_id']}) "
                       f"{old_mn} → {membership}")
                print(f"  ✎  {msg}")
                actions_log.append(msg)
                stats["membership_corrected"] += 1
            else:
                print(f"  ≈  [por nome] '{name}' encontrado (UserId {user['user_id']}), "
                      f"membership já correto ({membership}).")

        # ── 3. Garantir AthleteProfile ────────────────────────────
        athlete_profile_id = user["athlete_profile_id"]
        if athlete_profile_id is None:
            athlete_profile_id = create_athlete_profile(cursor, conn, user["user_id"])
            msg = (f"[ATLETA CRIADO] '{name}' (UserId {user['user_id']}) "
                   f"→ AthleteProfile Id {athlete_profile_id}")
            print(f"  ➕  {msg}")
            actions_log.append(msg)
            stats["athlete_profile_created"] += 1

        # ── 4. Garantir atribuição à equipa ───────────────────────
        if is_athlete_in_team(cursor, athlete_profile_id, team_id):
            stats["membership_ok"] += 1
        else:
            assign_athlete_to_team(cursor, conn, athlete_profile_id, team_id)
            msg = (f"[EQUIPA ATRIBUÍDA] '{name}' (UserId {user['user_id']}) "
                   f"→ TeamId {team_id} ('{team_excel}')")
            print(f"  ✔  {msg}")
            actions_log.append(msg)
            stats["team_added"] += 1

    # ── Relatório final ───────────────────────────────────────────
    print("\n" + "═" * 60)
    print("  RELATÓRIO FINAL")
    print("═" * 60)
    print(f"  Atletas no Excel processados   : {len(athletes)}")
    print(f"  Já corretos (sem alterações)   : {stats['membership_ok']}")
    print(f"  Adicionados a equipa           : {stats['team_added']}")
    print(f"  AthleteProfile criado          : {stats['athlete_profile_created']}")
    print(f"  MembershipNumber corrigido     : {stats['membership_corrected']}")
    print(f"  Equipa não resolvida na BD     : {stats['team_not_found']}")
    print(f"  Utilizadores não encontrados   : {stats['not_found']}")
    print("═" * 60)

    if team_unresolved:
        print("\n⚠  EQUIPAS NÃO RESOLVIDAS (verifica os nomes na BD):")
        for r in team_unresolved:
            print(f"   - '{r['name']}' → equipa Excel: '{r['team_excel']}' (sócio {r['socio']})")

    if not_found_list:
        print("\n✘  UTILIZADORES NÃO ENCONTRADOS (nem por sócio nem por nome):")
        for r in not_found_list:
            print(f"   - '{r['name']}' (sócio {r['socio']}, equipa '{r['team_excel']}')")

    print()


# ─────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 60)
    print("  SYNC BASQUETEBOL — CDP")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # Lê o Excel
    athletes = read_excel(EXCEL_PATH)

    # Liga à BD
    print("A ligar à base de dados...")
    try:
        conn = pyodbc.connect(CONNECTION_STRING)
        print("Ligação estabelecida.\n")
    except Exception as e:
        print(f"\n❌ Erro ao ligar à BD: {e}")
        sys.exit(1)

    try:
        process(athletes, conn)
    finally:
        conn.close()
        print("Ligação fechada.")